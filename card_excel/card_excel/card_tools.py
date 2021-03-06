import openpyxl

card_list = []   # 空列表

def card_menu():
    """
    名片管理系统 主界面
    :return:
    """
    print('*' * 30)
    print('欢迎使用【名片管理系统 Excel】 V1.0')
    print()
    print('1. 新增名片')
    print('2. 显示全部名片')
    print('3. 搜索名片')
    print()
    print('0. 退出系统')
    print('*' * 30)


def card_add():
    """
    增加名片
    :param filename: 文件名
    :param sheet: 表单名
    :return:
    """
    print("请按提示输入内容：")
    card_name = input('请输入姓名: ')
    card_phone = input('请输入电话号码: ')
    card_QQ = input('请输入QQ: ')
    card_Email = input('请输入邮箱: ')
    card_user_info = {'name': card_name,
                      'phone': card_phone,
                      'QQ': card_QQ,
                      'Email': card_Email}
    card_list.append(card_user_info)
    print(card_user_info)
    print('添加 ', card_name, ' 名片成功')


def card_show():
    """
    显示存储的名片
    :param filename: 文件名
    :param sheet: 表单名
    :return:
    """
    if  len(card_list) == 0:
        print("没有存储名片 请使用新增功能")
        return

    print('姓名\t\t电话\t\tQQ\t\t邮箱')
    print('=' * 30)
    for card_dict in card_list:
        print('{}\t\t{}\t\t{}\t\t{}'.format(card_dict['name'],
                                            card_dict['phone'],
                                            card_dict['QQ'],
                                            card_dict['Email']))


def card_search():

    # 用户输入
    print('搜索名片')
    card_name = input("请输入你要搜索的名片姓名：")
    # 判断用户输入 查询

    for card_find in card_list:
        if card_name == card_find['name']:
            print('姓名\t\t电话\t\tQQ\t\t邮箱')
            print('=' * 30)
            print('{}\t\t{}\t\t{}\t\t{}'.format(card_find['name'],
                                                      card_find['phone'],
                                                      card_find['QQ'],
                                                      card_find['Email']))

            # 出来查询到的名片
            card_deal_find(card_find)
            break
    else:
        print('没有找到名片您想找的名片')


def card_deal_find(card_find):
    print(card_find)
    card_use = input('请选择要执行的操作 [1]修改 [2]删除 [0]返回上级')
    if card_use == '1':
        card_find['name'] = card_input_info(card_find['name'], '姓名:')
        card_find['phone'] = card_input_info(card_find['phone'], '电话：')
        card_find['QQ'] = card_input_info(card_find['QQ'], 'qq：')
        card_find['Email'] = card_input_info(card_find['Email'], '邮箱：')

    elif card_use == '2':
        card_list.remove(card_find)
        print('已删除 %s 的名片' % card_find['name'])
        print(card_list)


def card_input_info(find_value, info):
    # 提示用户输入信息
    user_info = input(info)

    # 判断 用户有输入
    if len(user_info) > 0:
        return user_info
    # 用户没有输入
    else:
        return find_value


def card_excel_open(filename, sheet):
    global card_list    # 加global关键字 标记为全局变量
    card_list = []
    wb = openpyxl.load_workbook(filename)  # 加载excel文件
    sheet = wb[sheet]   # 找到表单
    max_row = sheet.max_row  # 获取最大行数
    for i in range(1, max_row+1):
        cell = sheet.cell(row=i, column=1).value  # 接收单元格的字典内容
        if not cell is None:
            # print(cell)
            # print(type(cell))
            cell = eval(cell)
            card_list.append(cell)


def card_excel_write(filename, sheet, write):
    """
    对文件进行写入
    :param filename: 文件名
    :param sheet: 表单名
    :param write: 写入名片列表
    :return:
    """
    if len(card_list) != 0:
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheet]
        i = 0
        for card_dict in write:
            i += 1
            sheet.cell(i, 1).value = str(card_dict)  # 找到特定单元格，进行写入

        wb.save(filename)   # 保存excel


def card_excel_delete(filename, sheet):
    """
    自己获取文件最大行，自己对文件进行整个删除
    :param filename: 文件名
    :param sheet: 表单名
    :return:
    """
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet]
    max_row = sheet.max_row
    sheet.delete_rows(1, max_row)  # 删除行内容
    wb.save(filename)

def card_excel_rewrite(filename, sheet, write):
    card_excel_delete(filename, sheet)
    card_excel_write(filename, sheet, write)


# 测试
# card_excel_delete('card_file.xlsx', 'Sheet1')
# write_repeat('card_file.xlsx', 'Sheet1', [{'name': '2', 'phone': '2', 'QQ': '2', 'Email': '2'}])