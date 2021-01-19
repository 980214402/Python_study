
# import openpyxl
#
# wb = openpyxl.load_workbook('test.xlsx') # 加载excel
# sheet1 = wb['Sheet1']  # 找到表单
# cell1 = sheet1.cell(row=1, column=1) # 通过行和列找到单元格得 key
# print(cell1)    # 打印单元格key
# print(cell1.value)  # 打印单元格的value
# cell1.value = "Hello Word"  # 写入单元格内容
# print(cell1.value)
# wb.save('test.xlsx')    # 保存文件

# import openpyxl
#
# # wb = openpyxl.Workbook('test1.xlsx')
# # wb.save('test1.xlsx')
# wb = openpyxl.load_workbook('test.xlsx')
# sheet1 = wb['Sheet1']
#
# cell1 = sheet1.cell(row=1, column=1)
# print(cell1)
# print(cell1.value)
# cell1.value = 'Hello'
# print(cell1.value)
# wb.save('test.xlsx')

import openpyxl
import requests

wb = openpyxl.load_workbook('test.xlsx')
sheet1 = wb['Sheet1']

url = sheet1.cell(row=2, column=5).value
data = eval(sheet1.cell(row=2, column=6).value)
expected_result = sheet1.cell(row=2, column=7).value
expected_result = expected_result.replace('null', 'None')
expected_result = eval(expected_result)


response = requests.post(url=url, data=data)
sheet1.cell(row=2, column=8).value = response.text

response = response.json()

if response['msg'] == expected_result['msg']:
    print('通过')
    sheet1.cell(row=2, column=9).value = '通过'
else:
    print('不通过')
    sheet1.cell(row=2, column=9).value = '不通过'

wb.save('test.xlsx')