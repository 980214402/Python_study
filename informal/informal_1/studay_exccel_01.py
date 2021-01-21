"""
data = eval(data)  赋值给data变量
    eval() 内置函数 把像字典格式的字符串转化为字典

expected_result = expected_result.replace('null', 'None') # 把null 替换成None
    .replace('a', 'b') 把字符串中的'a' 转化为 'b'
"""

import openpyxl
import requests

def read_data(filename, sheetname):
    """
    打开excel，读取测试用例
    :param filename: excel的文件名
    :param sheetname: excel中的表单名
    :return: 本表单中的所有测试用例
    """
    wb = openpyxl.load_workbook(filename) # 加载excel工作簿对象
    sheet = wb[sheetname] # 找到表单
    max_row = sheet.max_row  # 获取最大行数
    max_column = sheet.max_column # 获取最大列数
    case_list = []
    for i in range(2, max_row+1):
        case = dict(
            case_id = sheet.cell(row=i, column=1).value,
            url = sheet.cell(row=i, column=5).value,
            data = sheet.cell(row=i, column=6).value,
            expected_result = sheet.cell(row=i, column=7).value
        )
        case_list.append(case) # 每次循环往列表追加一个元素

    return case_list # 装了所有测试用例的列表


def post_func(post_url, post_data):
    """
    根据测试用例对接口post请求，并返回测试结果
    :param post_url: 接口地址
    :param post_data: 发送的post请求
    :return: 返回接口结果
    """
    response = requests.post(url=post_url, data=post_data)
    html = response.json() # 获取返回结果的字典格式

    return html  # 返回测试的结果


def write_result(filename, sheetname, row, column, final_result):
    """
    对测试结果进行回写
    :param filename: 同一目录下excel的文件名
    :param sheetname: excel中的表单ming
    :param row: 行
    :param column: 列
    :param final_result: 写入内容
    :return:
    """
    wb = openpyxl.load_workbook(filename)  # 加载工作簿
    sheet = wb[sheetname]  # 找到表单
    sheet.cell(row, column).value = final_result # 结果回写
    wb.save(filename)  # 保存文件


def execute_func(filename, sheetname):
    """
    接口实战，调用函数，整合
    :param filename: 同一目录下的excel文件名
    :param sheetname: excel中的表单名
    :return:
    """
    cases = read_data(filename, sheetname)  # 调用read_data函数
    for case in cases:
        case_id = case['case_id']
        url = case['url']
        data = case['data']  # 取出来是 字符串
        data = eval(data)
        expected_result = case.get('expected_result')  # 期望结果 字符串
        expected_result = expected_result.replace('null', 'None') # eval() 函数不识别null
        expected_result = eval(expected_result)
        expected_msg = expected_result['msg']  # 期望的msg结果

        real_result = post_func(url, data) # 调用post方法 发送http接口请求
        real_msg = real_result['msg']  # 实际的msg结果
        write_result(filename, sheetname, case_id+1, 8, str(real_result)) # 把实际返回结果回写

        if real_msg == expected_msg:
            print('第{}条测试用例执行通过！'.format(case_id))
            final_result = 'passed'
        else:
            print('第{}条测试用例执行不通过'.format(case_id))
            final_result = 'Failed'
        print('-' * 10)

        write_result(filename, sheetname, case_id+1, 9, final_result) # 调用写入函数，把对比结果写入


execute_func('test.xlsx', 'Sheet1')
